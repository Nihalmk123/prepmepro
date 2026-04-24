import frappe
import json
import openpyxl
import pdfplumber
from datetime import datetime


@frappe.whitelist(allow_guest=True)
def upload_and_parse(file_url=None, file_type=None):
    file_url = file_url or frappe.form_dict.get("file_url")
    file_type = (file_type or frappe.form_dict.get("file_type") or "").strip().lower()

    if not file_url:
        frappe.throw("file_url is required")
    if file_type not in ("excel", "pdf"):
        frappe.throw("file_type must be excel or pdf")

    files = frappe.get_all(
        "File",
        filters={"file_url": file_url},
        fields=["name", "file_url"],
        limit=1
    )
    if not files:
        frappe.throw(f"File not found for url: {file_url}")

    file_doc = frappe.get_doc("File", files[0].name)
    file_path = file_doc.get_full_path()

    frappe.db.delete("Content Record", {"source_file": file_url})

    extracted = []

    if file_type == "excel":
        count, extracted = _parse_excel(file_path, file_url)
    else:
        count, extracted = _parse_pdf(file_path, file_url)

    frappe.db.commit()

    return {
        "status": "success",
        "records_inserted": count,
        "source": file_url,
        "file_type": file_type,
        "data": extracted
    }


def _parse_excel(file_path, source_file):
    wb = openpyxl.load_workbook(file_path, data_only=True)
    ws = wb.active
    headers = [str(c.value).strip() if c.value else f"col_{i+1}" for i, c in enumerate(ws[1])]

    count = 0
    extracted = []

    for idx, row in enumerate(ws.iter_rows(min_row=2, values_only=True), start=2):
        if all(v is None for v in row):
            continue

        row_dict = dict(zip(headers, row))
        title = (
            row_dict.get("title") or row_dict.get("Title") or
            row_dict.get("name") or row_dict.get("Name") or
            f"Row {idx}"
        )

        payload = {
            "row_index": idx,
            "title": str(title),
            "description": str(row_dict.get("description") or row_dict.get("Description") or ""),
            "category": str(row_dict.get("category") or row_dict.get("Category") or ""),
            "data": row_dict
        }

        doc = frappe.get_doc({
            "doctype": "Content Record",
            "source_file": source_file,
            "file_type": "excel",
            "row_index": idx,
            "title": payload["title"],
            "description": payload["description"],
            "category": payload["category"],
            "raw_json": json.dumps(row_dict, default=str),
            "uploaded_on": datetime.now()
        })
        doc.insert(ignore_permissions=True)

        extracted.append(payload)
        count += 1

    return count, extracted


def _parse_pdf(file_path, source_file):
    count = 0
    extracted = []

    with pdfplumber.open(file_path) as pdf:
        for idx, page in enumerate(pdf.pages, start=1):
            text = page.extract_text() or ""
            tables = page.extract_tables() or []

            payload = {
                "page": idx,
                "text": text,
                "tables": tables
            }

            doc = frappe.get_doc({
                "doctype": "Content Record",
                "source_file": source_file,
                "file_type": "pdf",
                "row_index": idx,
                "title": f"Page {idx}",
                "description": text,
                "category": "",
                "raw_json": json.dumps(payload, default=str),
                "uploaded_on": datetime.now()
            })
            doc.insert(ignore_permissions=True)

            extracted.append(payload)
            count += 1

    return count, extracted
